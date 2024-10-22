import pandas as pd
import fitz  # PyMuPDF
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import os

def extract_text_and_images_from_pdf(pdf_path):
    """Extract text and images from the PDF."""
    # Open the PDF file
    pdf_document = fitz.open(pdf_path)
    data = []

    for page_number in range(len(pdf_document)):
        page = pdf_document.load_page(page_number)
        
        # Extract text
        text = page.get_text("text")

        # Extract images
        image_list = page.get_images(full=True)
        images = []
        for img_index, img_info in enumerate(image_list):
            xref = img_info[0]
            base_image = pdf_document.extract_image(xref)
            image_bytes = base_image["image"]

            # Convert bytes to an image
            img = Image.open(BytesIO(image_bytes))
            images.append(img)

        # Append extracted data for this page
        data.append({
            "page_number": page_number + 1,
            "text": text,
            "images": images
        })

    pdf_document.close()
    return data
'''
def save_to_excel(data, excel_path):
    """Save extracted text and images into an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Data"

    # Adding header
    ws['A1'] = 'Page Number'
    ws['B1'] = 'Text'
    ws['C1'] = 'Images'

    # Initialize pandas DataFrame to format text data
    text_data = []
    image_counter = 1
    image_dir = "temp_images"
    os.makedirs(image_dir, exist_ok=True)  # Create directory for temp images

    # Writing data into the Excel file
    for row_index, item in enumerate(data, start=2):
        page_number = item['page_number']
        text = item['text']
        images = item['images']

        ws[f'A{row_index}'] = page_number
        ws[f'B{row_index}'] = text

        if images:
            for img in images:
                # Save the image temporarily
                img_name = f"{image_dir}/image_{page_number}_{image_counter}.png"
                img.save(img_name)

                # Load the image into the Excel file
                excel_image = ExcelImage(img_name)
                img_cell = f'C{row_index}'
                ws.add_image(excel_image, img_cell)

                image_counter += 1

        text_data.append({
            "Page Number": page_number,
            "Text": text
        })

    # Save workbook
    wb.save(excel_path)

    # Clean up temporary images
    for img_file in os.listdir(image_dir):
        os.remove(os.path.join(image_dir, img_file))
    os.rmdir(image_dir)
'''
def save_to_excel(data, excel_path):
    """Save extracted text and images into an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Data"

    # Adding header
    ws['A1'] = 'Page Number'
    ws['B1'] = 'Text'
    ws['C1'] = 'Images'

    # Set column widths to accommodate images and text
    ws.column_dimensions['B'].width = 50  # Widen column for text
    ws.column_dimensions['C'].width = 20  # Widen column for images

    # Initialize pandas DataFrame to format text data
    image_counter = 1
    image_dir = "temp_images"
    os.makedirs(image_dir, exist_ok=True)  # Create directory for temp images

    # Writing data into the Excel file
    for row_index, item in enumerate(data, start=2):
        page_number = item['page_number']
        text = item['text']
        images = item['images']

        # Insert text
        ws[f'A{row_index}'] = page_number
        ws[f'B{row_index}'] = text

        # Insert images into a specific cell
        if images:
            for img in images:
                # Save the image temporarily
                img_name = f"{image_dir}/image_{page_number}_{image_counter}.png"
                img.save(img_name)

                # Load the image into Excel
                img_for_excel = ExcelImage(img_name)

                # Resize image to fit in the cell (optional, adjust to your needs)
                img_for_excel.width = 100  # adjust width
                img_for_excel.height = 100  # adjust height

                # Insert image into a specific cell (column 'C')
                ws.add_image(img_for_excel, f'C{row_index}')

                image_counter += 1

    # Save the workbook
    wb.save(excel_path)

    # Clean up temporary images
    for img_file in os.listdir(image_dir):
        os.remove(os.path.join(image_dir, img_file))
    os.rmdir(image_dir)


def main():
    # Specify the paths for the PDF and Excel file
    pdf_path = "input.pdf"
    excel_path = "output.xlsx"

    # Extract data from PDF
    data = extract_text_and_images_from_pdf(pdf_path)

    # Save the extracted data to Excel
    save_to_excel(data, excel_path)

    print(f"Data has been successfully extracted and saved to {excel_path}")

if __name__ == "__main__":
    main()
